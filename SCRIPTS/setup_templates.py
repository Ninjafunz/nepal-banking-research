"""
setup_templates.py
==================
Creates all 10 Excel data-entry templates (01 to 10) plus the ANALYSIS template.
Run once to scaffold the project. Each file has pre-populated headers, a
colour-coded header row, bank_code / fy keys in every row, and data-validation
drop-downs where applicable.

Usage:
    cd SCRIPTS
    python setup_templates.py
"""

import sys
import os
import glob

# Add SCRIPTS dir to path so banks.py is importable when run from any directory
sys.path.insert(0, os.path.dirname(__file__))

# Windows Store Python stores user-installed packages in a non-standard location.
# Inject all known candidate paths so openpyxl/pandas/etc. are always found.
_extra_paths = glob.glob(
    os.path.expanduser(
        r"~\AppData\Local\Packages\PythonSoftwareFoundation.Python.3*"
        r"\LocalCache\local-packages\Python3*\site-packages"
    )
) + glob.glob(
    os.path.expanduser(r"~\AppData\Roaming\Python\Python3*\site-packages")
)
for _p in _extra_paths:
    if _p not in sys.path:
        sys.path.insert(0, _p)

from banks import BANKS, BANK_CODES, BANK_MAP, FISCAL_YEARS, MACRO_YEARS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
KEY_FILL   = PatternFill("solid", fgColor="D6E4F0")   # light blue for key columns
AUTO_FILL  = PatternFill("solid", fgColor="E2EFDA")   # light green for auto-calculated
NOTE_FILL  = PatternFill("solid", fgColor="FFF2CC")   # yellow for source/notes
THIN       = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, headers, auto_cols=None, note_cols=None):
    """Apply header styles and set column widths."""
    auto_cols  = auto_cols  or []
    note_cols  = note_cols  or []
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)

def populate_keys(ws, keys_list, start_col=1):
    """Fill bank_code, bank_name, fy rows."""
    for row_idx, (bank_code, fy) in enumerate(keys_list, 2):
        ws.cell(row=row_idx, column=start_col,     value=bank_code).fill = KEY_FILL
        ws.cell(row=row_idx, column=start_col + 1, value=BANK_MAP.get(bank_code, "")).fill = KEY_FILL
        ws.cell(row=row_idx, column=start_col + 2, value=fy).fill = KEY_FILL

def freeze_and_autofilter(ws, freeze_col=3):
    """Freeze first N columns and add autofilter."""
    ws.freeze_panes = ws.cell(row=2, column=freeze_col + 1)
    ws.auto_filter.ref = ws.dimensions

def add_source_validation(ws, col_idx, start_row, end_row):
    dv = DataValidation(
        type="list",
        formula1='"NRB,AnnualReport,Calculated,NEPSE,Estimated"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv)
    for row in range(start_row, end_row + 1):
        dv.add(ws.cell(row=row, column=col_idx))

def save_wb(wb, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"  Created: {os.path.basename(path)}")


# ---------------------------------------------------------------------------
# Key pairs: (bank_code, fy) for bank-level datasets
# ---------------------------------------------------------------------------
BANK_FY_PAIRS = [(b["code"], fy) for b in BANKS for fy in FISCAL_YEARS]
N = len(BANK_FY_PAIRS)   # 23 banks x 6 years = 138 rows


# ===========================================================================
# 01 — Bank Financials
# ===========================================================================
def create_01_bank_financials(data_dir):
    wb = openpyxl.Workbook()

    # --- Sheet 1: Balance Sheet ---
    ws_bs = wb.active
    ws_bs.title = "balance_sheet"
    bs_headers = [
        "bank_code", "bank_name", "fy",
        "total_assets", "cash_bank_balances", "investments",
        "gross_loans", "net_loans",
        "total_deposits", "borrowings", "total_liabilities",
        "shareholders_equity", "paid_up_capital", "reserves",
        "source", "notes"
    ]
    style_header(ws_bs, bs_headers)
    populate_keys(ws_bs, BANK_FY_PAIRS)
    add_source_validation(ws_bs, 15, 2, N + 1)
    freeze_and_autofilter(ws_bs)

    # --- Sheet 2: Income Statement ---
    ws_is = wb.create_sheet("income_statement")
    is_headers = [
        "bank_code", "bank_name", "fy",
        "interest_income", "interest_expense", "net_interest_income",
        "non_interest_income", "operating_income",
        "operating_expenses", "personnel_expenses",
        "provision_loan_losses",
        "profit_before_tax", "profit_after_tax",
        "source", "notes"
    ]
    style_header(ws_is, is_headers)
    populate_keys(ws_is, BANK_FY_PAIRS)
    add_source_validation(ws_is, 14, 2, N + 1)
    freeze_and_autofilter(ws_is)

    # --- Sheet 3: Data Dictionary ---
    ws_dict = wb.create_sheet("data_dictionary")
    ws_dict.append(["Field", "Description", "Unit", "Source Priority"])
    dict_rows = [
        ("bank_code",          "NRB/NEPSE standard bank ticker",      "—",          "—"),
        ("bank_name",          "Full registered name",                "—",          "—"),
        ("fy",                 "Fiscal year end (Gregorian)",         "Year",       "—"),
        ("total_assets",       "Total balance sheet assets",          "NPR millions","NRB > AnnualReport"),
        ("cash_bank_balances", "Cash + balances with NRB & banks",   "NPR millions","NRB > AnnualReport"),
        ("investments",        "Securities + government bonds",       "NPR millions","NRB > AnnualReport"),
        ("gross_loans",        "Loans before provisions",            "NPR millions","NRB > AnnualReport"),
        ("net_loans",          "Loans after provisions",             "NPR millions","AnnualReport"),
        ("total_deposits",     "All customer deposits",              "NPR millions","NRB > AnnualReport"),
        ("borrowings",         "Inter-bank + institutional borrowing","NPR millions","AnnualReport"),
        ("total_liabilities",  "Total liabilities",                  "NPR millions","AnnualReport"),
        ("shareholders_equity","Book equity",                         "NPR millions","NRB > AnnualReport"),
        ("paid_up_capital",    "Issued and paid-up share capital",   "NPR millions","NRB > AnnualReport"),
        ("reserves",           "Retained earnings + reserves",       "NPR millions","AnnualReport"),
        ("interest_income",    "Total interest earned",              "NPR millions","NRB > AnnualReport"),
        ("interest_expense",   "Total interest paid",                "NPR millions","NRB > AnnualReport"),
        ("net_interest_income","interest_income - interest_expense", "NPR millions","Calculated"),
        ("non_interest_income","Fees, commissions, FX, trading",    "NPR millions","AnnualReport"),
        ("operating_income",   "NII + non_interest_income",         "NPR millions","Calculated"),
        ("operating_expenses", "Total non-interest expense",        "NPR millions","AnnualReport"),
        ("personnel_expenses", "Salaries, allowances, benefits",    "NPR millions","AnnualReport"),
        ("provision_loan_losses","Loan loss provision charge",       "NPR millions","AnnualReport"),
        ("profit_before_tax",  "PBT",                               "NPR millions","NRB > AnnualReport"),
        ("profit_after_tax",   "Net profit (PAT)",                  "NPR millions","NRB > AnnualReport"),
    ]
    for row in dict_rows:
        ws_dict.append(row)
    ws_dict.column_dimensions["A"].width = 25
    ws_dict.column_dimensions["B"].width = 45
    ws_dict.column_dimensions["C"].width = 16
    ws_dict.column_dimensions["D"].width = 25

    save_wb(wb, os.path.join(data_dir, "01_bank_financials.xlsx"))


# ===========================================================================
# 02 — Bank Ratios (auto-generated shell only)
# ===========================================================================
def create_02_bank_ratios(data_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ratios"
    headers = [
        "bank_code", "bank_name", "fy",
        # Profitability
        "roa", "roe", "nim", "profit_margin",
        # Growth (YoY %)
        "asset_growth", "loan_growth", "deposit_growth", "equity_growth", "profit_growth",
        # Efficiency
        "cost_income",
        "assets_per_employee", "loans_per_employee", "deposits_per_employee", "profit_per_employee",
        "assets_per_branch",
        # Risk
        "gross_npl_pct", "net_npl_pct", "provision_coverage_pct", "cost_of_risk",
        "car_pct", "cet1_pct",
        # Liquidity/Funding
        "loan_deposit_ratio", "casa_ratio", "fixed_deposit_share", "cost_of_deposits_pct",
        # Notes
        "calc_notes"
    ]
    style_header(ws, headers, auto_cols=list(range(4, len(headers) + 1)))
    populate_keys(ws, BANK_FY_PAIRS)
    freeze_and_autofilter(ws)

    # Mark all data columns as auto-calculated (green tint)
    for row in range(2, N + 2):
        for col in range(4, len(headers)):
            ws.cell(row=row, column=col).fill = AUTO_FILL

    ws_note = wb.create_sheet("notes")
    ws_note["A1"] = "AUTO-GENERATED FILE"
    ws_note["A2"] = "Do not edit manually. Run SCRIPTS/calculate_ratios.py to regenerate."
    ws_note["A3"] = "All ratio fields use NPR millions as base unit."
    ws_note["A4"] = "ROA / ROE / NIM use average-balance convention ((t + t-1) / 2)."

    save_wb(wb, os.path.join(data_dir, "02_bank_ratios.xlsx"))


# ===========================================================================
# 03 — Market Shares (auto-generated shell only)
# ===========================================================================
def create_03_market_shares(data_dir):
    wb = openpyxl.Workbook()

    # Sheet 1: Per-bank market shares
    ws_ms = wb.active
    ws_ms.title = "market_shares"
    ms_headers = [
        "bank_code", "bank_name", "fy",
        "asset_share_pct", "loan_share_pct", "deposit_share_pct", "profit_share_pct",
        "asset_rank", "loan_rank", "deposit_rank", "profit_rank"
    ]
    style_header(ws_ms, ms_headers)
    populate_keys(ws_ms, BANK_FY_PAIRS)
    for row in range(2, N + 2):
        for col in range(4, len(ms_headers) + 1):
            ws_ms.cell(row=row, column=col).fill = AUTO_FILL
    freeze_and_autofilter(ws_ms)

    # Sheet 2: Industry concentration
    ws_conc = wb.create_sheet("concentration")
    conc_headers = [
        "fy",
        "hhi_assets", "hhi_loans", "hhi_deposits",
        "cr4_assets",  "cr5_assets",  "cr10_assets",
        "cr4_loans",   "cr5_loans",   "cr10_loans",
        "cr4_deposits","cr5_deposits","cr10_deposits",
        "n_banks",
        "interpretation"
    ]
    style_header(ws_conc, conc_headers)
    for row_idx, yr in enumerate(FISCAL_YEARS, 2):
        ws_conc.cell(row=row_idx, column=1, value=yr).fill = KEY_FILL
        for col in range(2, len(conc_headers) + 1):
            ws_conc.cell(row=row_idx, column=col).fill = AUTO_FILL
    ws_conc.freeze_panes = "B2"

    ws_note = wb.create_sheet("notes")
    ws_note["A1"] = "AUTO-GENERATED FILE"
    ws_note["A2"] = "Run SCRIPTS/calculate_market_shares.py to regenerate."
    ws_note["A3"] = "HHI uses percentage-squared convention (max = 10,000)."
    ws_note["A4"] = "CR4/CR5/CR10 = cumulative share of N largest banks."

    save_wb(wb, os.path.join(data_dir, "03_market_shares.xlsx"))


# ===========================================================================
# 04 — Operating Metrics
# ===========================================================================
def create_04_operating_metrics(data_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "operating_metrics"
    headers = [
        "bank_code", "bank_name", "fy",
        # Physical network
        "branches", "employees", "atms", "extension_counters", "branchless_banking_centers",
        # Digital channels
        "mobile_banking_users", "internet_banking_users",
        "debit_cards", "credit_cards", "qr_users",
        "digital_transactions_count", "agent_network_points",
        # Risk metrics (also in 02, sourced here)
        "gross_npl_pct", "net_npl_pct", "provision_coverage_pct",
        "car_pct", "cet1_pct",
        "source", "notes"
    ]
    style_header(ws, headers)
    populate_keys(ws, BANK_FY_PAIRS)
    add_source_validation(ws, headers.index("source") + 1, 2, N + 1)
    freeze_and_autofilter(ws)
    save_wb(wb, os.path.join(data_dir, "04_operating_metrics.xlsx"))


# ===========================================================================
# 05 — Loan Composition
# ===========================================================================
def create_05_loan_composition(data_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "loan_composition"
    headers = [
        "bank_code", "bank_name", "fy",
        # NRB sectoral categories
        "agriculture", "manufacturing", "construction",
        "wholesale_retail", "transportation", "tourism",
        "consumption", "real_estate", "hydropower",
        # Segment breakdown (where available)
        "sme", "retail", "corporate", "housing",
        "vehicle", "margin_lending", "other_sectors",
        # Validation
        "total_loans_check",
        "source", "notes"
    ]
    style_header(ws, headers)
    populate_keys(ws, BANK_FY_PAIRS)
    # total_loans_check is auto (green)
    tc_col = headers.index("total_loans_check") + 1
    for row in range(2, N + 2):
        ws.cell(row=row, column=tc_col).fill = AUTO_FILL
    add_source_validation(ws, headers.index("source") + 1, 2, N + 1)
    freeze_and_autofilter(ws)
    save_wb(wb, os.path.join(data_dir, "05_loan_composition.xlsx"))


# ===========================================================================
# 06 — Deposit Composition
# ===========================================================================
def create_06_deposit_composition(data_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "deposit_composition"
    headers = [
        "bank_code", "bank_name", "fy",
        "current_deposits", "savings_deposits", "fixed_deposits",
        "call_deposits", "other_deposits",
        # Calculated
        "total_deposits_check", "casa_ratio", "fixed_deposit_share",
        "cost_of_deposits_pct",
        "source", "notes"
    ]
    style_header(ws, headers)
    populate_keys(ws, BANK_FY_PAIRS)
    auto_start = headers.index("total_deposits_check") + 1
    auto_end   = headers.index("cost_of_deposits_pct") + 1
    for row in range(2, N + 2):
        for col in range(auto_start, auto_end + 1):
            ws.cell(row=row, column=col).fill = AUTO_FILL
    add_source_validation(ws, headers.index("source") + 1, 2, N + 1)
    freeze_and_autofilter(ws)
    save_wb(wb, os.path.join(data_dir, "06_deposit_composition.xlsx"))


# ===========================================================================
# 07 — Macro Indicators
# ===========================================================================
def create_07_macro_indicators(data_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "macro_indicators"
    headers = [
        "fy",
        # Macroeconomy
        "gdp_growth_pct", "inflation_pct",
        "remittance_growth_pct", "remittance_usd_mn",
        "private_credit_growth_pct",
        "government_spending_growth_pct",
        # Monetary
        "policy_rate_pct", "bank_rate_pct",
        "interbank_rate_pct", "avg_deposit_rate_pct", "avg_lending_rate_pct",
        "liquidity_ratio_pct",
        # Banking system
        "system_total_assets", "system_total_loans", "system_total_deposits",
        "system_loan_growth_pct", "system_deposit_growth_pct",
        "system_npl_pct", "system_roe_pct", "system_nim_pct", "system_car_pct",
        "n_commercial_banks",
        "source", "notes"
    ]
    style_header(ws, headers)
    for row_idx, yr in enumerate(MACRO_YEARS, 2):
        ws.cell(row=row_idx, column=1, value=yr).fill = KEY_FILL
    add_source_validation(ws, headers.index("source") + 1, 2, len(MACRO_YEARS) + 1)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    save_wb(wb, os.path.join(data_dir, "07_macro_indicators.xlsx"))


# ===========================================================================
# 08 — Market Data (listed banks)
# ===========================================================================
def create_08_market_data(data_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "market_data"
    listed_pairs = [
        (b["code"], yr)
        for b in BANKS if b["listed"]
        for yr in MACRO_YEARS
    ]
    headers = [
        "bank_code", "bank_name", "fy",
        "ticker",
        "share_price_eoy", "market_cap",
        "pe_ratio", "pb_ratio",
        "eps", "bvps",
        "dividend_per_share", "dividend_yield_pct",
        "annual_return_pct", "price_volatility",
        "shares_outstanding",
        "source", "notes"
    ]
    style_header(ws, headers)
    for row_idx, (code, yr) in enumerate(listed_pairs, 2):
        ws.cell(row=row_idx, column=1, value=code).fill = KEY_FILL
        ws.cell(row=row_idx, column=2, value=BANK_MAP.get(code, "")).fill = KEY_FILL
        ws.cell(row=row_idx, column=3, value=yr).fill = KEY_FILL
    add_source_validation(ws, headers.index("source") + 1, 2, len(listed_pairs) + 1)
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    save_wb(wb, os.path.join(data_dir, "08_market_data.xlsx"))


# ===========================================================================
# 09 — Strategic Coding
# ===========================================================================
def create_09_strategic_coding(data_dir):
    wb = openpyxl.Workbook()

    # Sheet 1: Digital scorecard
    ws_dig = wb.active
    ws_dig.title = "digital_scorecard"
    dig_headers = [
        "bank_code", "bank_name", "fy",
        "digital_account_opening", "mobile_banking", "digital_lending",
        "qr_ecosystem", "api_open_banking", "ai_initiatives",
        "digital_customer_acquisition", "core_banking_upgrade",
        "fintech_partnership", "cybersecurity_initiative",
        "digital_index",
        "evidence_notes"
    ]
    style_header(ws_dig, dig_headers)
    populate_keys(ws_dig, BANK_FY_PAIRS)
    di_col = dig_headers.index("digital_index") + 1
    for row in range(2, N + 2):
        ws_dig.cell(row=row, column=di_col).fill = AUTO_FILL
    # 0/1 validation for indicator columns
    dv_binary = DataValidation(type="list", formula1='"0,1"', allow_blank=True, showDropDown=False)
    ws_dig.add_data_validation(dv_binary)
    for col in range(4, di_col):
        for row in range(2, N + 2):
            dv_binary.add(ws_dig.cell(row=row, column=col))
    freeze_and_autofilter(ws_dig)

    # Sheet 2: Strategic priorities
    ws_strat = wb.create_sheet("strategic_priorities")
    strat_headers = [
        "bank_code", "bank_name", "fy",
        "priority_retail", "priority_sme", "priority_corporate",
        "priority_digital", "priority_branch_expansion", "priority_cost_reduction",
        "priority_wealth_mgmt", "priority_remittance", "priority_sustainability",
        "priority_geographic_expansion",
        "strategic_score",
        "evidence_notes"
    ]
    style_header(ws_strat, strat_headers)
    populate_keys(ws_strat, BANK_FY_PAIRS)
    ss_col = strat_headers.index("strategic_score") + 1
    for row in range(2, N + 2):
        ws_strat.cell(row=row, column=ss_col).fill = AUTO_FILL
    dv_binary2 = DataValidation(type="list", formula1='"0,1"', allow_blank=True, showDropDown=False)
    ws_strat.add_data_validation(dv_binary2)
    for col in range(4, ss_col):
        for row in range(2, N + 2):
            dv_binary2.add(ws_strat.cell(row=row, column=col))
    freeze_and_autofilter(ws_strat)

    # Sheet 3: Coding methodology
    ws_meth = wb.create_sheet("methodology")
    rows = [
        ["CODING METHODOLOGY — 09_strategic_coding.xlsx"],
        [],
        ["PRINCIPLE: All codes must be evidence-based. Cite the source in evidence_notes."],
        [],
        ["Digital Scorecard (0 = absent, 1 = present/active in that FY)"],
        ["Indicator",              "Code 1 if...",                                            "Typical source"],
        ["digital_account_opening","Bank offers fully digital account opening",               "Annual report, website"],
        ["mobile_banking",         "Mobile banking app active with core transactions",        "Annual report, NRB stats"],
        ["digital_lending",        "Digital/online loan products available",                  "Annual report"],
        ["qr_ecosystem",           "QR payment active (merchant or P2P)",                    "Annual report, NRB"],
        ["api_open_banking",       "Published API or open-banking integration",               "Annual report, press"],
        ["ai_initiatives",         "Named AI/ML project in production or development",        "Annual report"],
        ["digital_customer_acquisition","Online/app-based customer onboarding active",        "Annual report"],
        ["core_banking_upgrade",   "Major core banking system migration in year",             "Annual report"],
        ["fintech_partnership",    "Named fintech/technology partner agreement",              "Annual report, press"],
        ["cybersecurity_initiative","Named cybersecurity investment or certification",        "Annual report"],
        [],
        ["DO NOT inflate scores. If evidence is ambiguous, code 0 and note the uncertainty."],
    ]
    for r in rows:
        ws_meth.append(r)
    ws_meth.column_dimensions["A"].width = 30
    ws_meth.column_dimensions["B"].width = 55
    ws_meth.column_dimensions["C"].width = 30

    save_wb(wb, os.path.join(data_dir, "09_strategic_coding.xlsx"))


# ===========================================================================
# 10 — Bank Events
# ===========================================================================
def create_10_bank_events(data_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "events"
    headers = [
        "event_id", "bank_code", "bank_name", "fy",
        "event_date", "event_type", "event_description",
        "strategic_impact", "counterparty",
        "financial_effect_notes",
        "source_url", "notes"
    ]
    style_header(ws, headers)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["G"].width = 50
    ws.column_dimensions["J"].width = 40
    ws.column_dimensions["K"].width = 40

    # Event type validation
    dv_type = DataValidation(
        type="list",
        formula1='"M&A,Technology,Regulatory,Capital Raise,Leadership,Restructuring,Product Launch,Other"',
        allow_blank=True, showDropDown=False
    )
    # Impact validation
    dv_impact = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True, showDropDown=False
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_impact)
    for row in range(2, 200):
        dv_type.add(ws.cell(row=row, column=6))
        dv_impact.add(ws.cell(row=row, column=8))
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = "A1:L1"
    save_wb(wb, os.path.join(data_dir, "10_bank_events.xlsx"))


# ===========================================================================
# ANALYSIS — Industry Structure
# ===========================================================================
def create_analysis_template(analysis_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "industry_structure"
    headers = [
        "fy",
        "hhi_assets", "hhi_loans", "hhi_deposits",
        "cr4_assets",  "cr5_assets",  "cr10_assets",
        "cr4_loans",   "cr5_loans",   "cr10_loans",
        "cr4_deposits","cr5_deposits","cr10_deposits",
        "n_banks", "interpretation"
    ]
    style_header(ws, headers)
    for row_idx, yr in enumerate(FISCAL_YEARS, 2):
        ws.cell(row=row_idx, column=1, value=yr).fill = KEY_FILL
        for col in range(2, len(headers)):
            ws.cell(row=row_idx, column=col).fill = AUTO_FILL
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    save_wb(wb, os.path.join(analysis_dir, "industry_structure.xlsx"))


# ===========================================================================
# Main
# ===========================================================================
if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir     = os.path.join(base, "DATA")
    analysis_dir = os.path.join(base, "ANALYSIS")

    print("Nepal Banking Research — Creating Excel templates...\n")
    create_01_bank_financials(data_dir)
    create_02_bank_ratios(data_dir)
    create_03_market_shares(data_dir)
    create_04_operating_metrics(data_dir)
    create_05_loan_composition(data_dir)
    create_06_deposit_composition(data_dir)
    create_07_macro_indicators(data_dir)
    create_08_market_data(data_dir)
    create_09_strategic_coding(data_dir)
    create_10_bank_events(data_dir)
    create_analysis_template(analysis_dir)

    print(f"\nDone. {11} files created.")
    print(f"  DATA dir:     {data_dir}")
    print(f"  ANALYSIS dir: {analysis_dir}")
    print("\nNext step: Run calculate_ratios.py and calculate_market_shares.py")
    print("after populating 01_bank_financials.xlsx and 04_operating_metrics.xlsx.")
